from traceback import print_tb
import openpyxl

inv_file = openpyxl.load_workbook("inventory.xlsx")
product_list = inv_file["Sheet1"]

product_per_supplier = {}
total_value_per_supplier = {}
products_under_10_inv = {}

# print(product_list.max_row)

# start iteration from row 2

for product_row in range(2, product_list.max_row + 1):
    supplier_name = product_list.cell(product_row, 4).value
    inventory = product_list.cell(product_row, 2).value
    price = product_list.cell(product_row, 3).value
    product_no = product_list.cell(product_row, 1).value
    inventory_price = product_list.cell(product_row, 5)

    # Task 1: Calculate how many product per supplier
    if supplier_name in product_per_supplier:
        current_no_prod = product_per_supplier.get(supplier_name)
        product_per_supplier[supplier_name] = current_no_prod + 1
    else:
        product_per_supplier[supplier_name] = 1

    # Task 2: Calc total inventory value of their product.
    if supplier_name in total_value_per_supplier:
        current_total_value = total_value_per_supplier.get(supplier_name)
        total_value_per_supplier[supplier_name] = (
            current_total_value + inventory * price
        )
    else:
        total_value_per_supplier[supplier_name] = inventory * price

    # logic - product with inventory less than 10
    if inventory < 10:
        products_under_10_inv[product_no] = inventory

    # add value for total inventory price
    inventory_price.value = inventory * price


inv_file.save("inventory_with_total_value.xlsx")
